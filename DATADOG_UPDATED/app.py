from flask import Flask, render_template, render_template_string, jsonify, send_file, request
import sqlite3
import io
import os
import random
import psutil
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── App setup ──────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND = os.path.join(BASE_DIR, "frontend")

app = Flask(__name__, static_folder=FRONTEND, static_url_path="/frontend")

DB_NAME = os.path.join(BASE_DIR, "reports.db")


# ══════════════════════════════════════════════════════
# DATABASE  (Reporting Module)
# ══════════════════════════════════════════════════════
def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def initialize_database():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS monitoring_reports (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            report_date   TEXT,
            service_name  TEXT,
            status        TEXT,
            uptime        REAL,
            response_time INTEGER,
            error_rate    REAL,
            alerts        INTEGER,
            cpu_usage     INTEGER,
            memory_usage  INTEGER,
            ai_insight    TEXT
        )
    """)
    cur.execute("SELECT COUNT(*) AS cnt FROM monitoring_reports")
    if cur.fetchone()["cnt"] == 0:
        sample = [
            ("2026-06-25","Payment API","Healthy",99.98,120,0.02,1,42,58,
             "AI detected stable payment traffic with no major anomalies."),
            ("2026-06-25","User Authentication","Warning",99.70,380,0.30,3,76,81,
             "AI detected higher login response time during peak traffic."),
            ("2026-06-25","Database Service","Healthy",99.99,90,0.01,0,38,62,
             "AI predicts normal database performance for the next 24 hours."),
            ("2026-06-24","Payment API","Healthy",99.97,135,0.03,1,46,60,
             "AI found minor traffic increase but system remains stable."),
            ("2026-06-24","User Authentication","Healthy",99.92,210,0.08,1,58,70,
             "AI detected improved login performance compared to yesterday."),
            ("2026-06-23","Database Service","Critical",98.80,850,1.20,7,91,95,
             "AI detected database overload and recommends scaling resources."),
            ("2026-06-20","Notification Service","Healthy",99.95,180,0.04,0,44,55,
             "AI predicts stable notification delivery performance."),
            ("2026-06-15","Analytics Engine","Warning",99.50,510,0.45,4,79,84,
             "AI found memory usage trend increasing over the past week."),
            ("2026-06-01","Payment API","Healthy",99.99,110,0.01,0,39,52,
             "AI confirms high reliability for payment service."),
        ]
        cur.executemany("""
            INSERT INTO monitoring_reports
            (report_date,service_name,status,uptime,response_time,error_rate,
             alerts,cpu_usage,memory_usage,ai_insight)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, sample)
    conn.commit()
    conn.close()


def get_report_data(report_type):
    conn = get_db_connection()
    cur  = conn.cursor()
    today = datetime.now().date()
    if report_type == "daily":
        start = today - timedelta(days=1);  title = "Daily Monitoring Report"
    elif report_type == "weekly":
        start = today - timedelta(days=7);  title = "Weekly Monitoring Report"
    else:
        start = today - timedelta(days=30); title = "Monthly Monitoring Report"

    cur.execute("""
        SELECT * FROM monitoring_reports
        WHERE report_date >= ? ORDER BY report_date DESC
    """, (start.strftime("%Y-%m-%d"),))
    rows = cur.fetchall()
    conn.close()
    reports = [dict(r) for r in rows]
    if not reports:
        reports = [{"id":0,"report_date":today.strftime("%Y-%m-%d"),
                    "service_name":"No Data Available","status":"Healthy",
                    "uptime":0,"response_time":0,"error_rate":0,"alerts":0,
                    "cpu_usage":0,"memory_usage":0,
                    "ai_insight":"No monitoring data found for this period."}]
    n = len(reports)
    summary = {
        "title":            title,
        "generated_on":     datetime.now().strftime("%d %B %Y, %I:%M %p"),
        "total_services":   n,
        "avg_uptime":       round(sum(r["uptime"] for r in reports) / n, 2),
        "total_alerts":     sum(r["alerts"] for r in reports),
        "avg_response":     round(sum(r["response_time"] for r in reports) / n, 2),
        "critical_services":sum(1 for r in reports if r["status"] == "Critical"),
    }
    return summary, reports


# ══════════════════════════════════════════════════════
# FRONTEND ROUTES  (serve HTML pages)
# ══════════════════════════════════════════════════════
@app.route("/")
@app.route("/index.html")
def root():
    return send_file(os.path.join(FRONTEND, "login.html"))


@app.route("/dashboard")
def dashboard():
    return send_file(os.path.join(FRONTEND, "index.html"))


# Serve any frontend module page: /module/<folder>/  or  /module/<folder>/index.html
@app.route("/module/<path:subpath>")
def module_page(subpath):
    if not subpath.endswith(".html"):
        subpath = subpath.rstrip("/") + "/index.html"
    target = os.path.join(FRONTEND, subpath)
    if os.path.exists(target):
        return send_file(target)
    return "Page not found", 404


# ══════════════════════════════════════════════════════
# REPORTING MODULE API
# ══════════════════════════════════════════════════════
@app.route("/api/report/<report_type>")
def api_report(report_type):
    if report_type not in ("daily","weekly","monthly"):
        return jsonify({"error": "Invalid report type"}), 400
    summary, reports = get_report_data(report_type)
    return jsonify({"summary": summary, "reports": reports})


@app.route("/export/pdf/<report_type>")
def export_pdf(report_type):
    summary, reports = get_report_data(report_type)
    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    # Header bar
    pdf.setFillColor(colors.HexColor("#101935"))
    pdf.rect(0, H-90, W, 90, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawString(40, H-48, "Datadog AI Observability")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(40, H-68, summary["title"])

    # Summary block
    pdf.setFillColor(colors.black)
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(40, H-120, "Report Summary")
    pdf.setFont("Helvetica", 10)
    for i, line in enumerate([
        f"Generated On : {summary['generated_on']}",
        f"Total Services: {summary['total_services']}",
        f"Average Uptime: {summary['avg_uptime']}%",
        f"Total Alerts  : {summary['total_alerts']}",
        f"Avg Response  : {summary['avg_response']} ms",
        f"Critical Svcs : {summary['critical_services']}",
    ]):
        pdf.drawString(40, H-145-i*18, line)

    # Table header
    y = H-290
    pdf.setFillColor(colors.HexColor("#3157d5"))
    pdf.rect(35, y, W-70, 24, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 9)
    for txt, x in zip(["Date","Service","Status","Uptime","Response","Alerts"],
                       [40,105,225,310,380,460]):
        pdf.drawString(x, y+8, txt)
    y -= 20

    pdf.setFont("Helvetica", 8)
    for r in reports:
        if y < 60:
            pdf.showPage(); y = H-50
        pdf.setFillColor(colors.black)
        pdf.drawString(40,  y, str(r["report_date"]))
        pdf.drawString(105, y, str(r["service_name"])[:20])
        pdf.drawString(225, y, str(r["status"]))
        pdf.drawString(310, y, f"{r['uptime']}%")
        pdf.drawString(380, y, f"{r['response_time']} ms")
        pdf.drawString(460, y, str(r["alerts"]))
        y -= 18

    pdf.setFont("Helvetica-Oblique", 8)
    pdf.setFillColor(colors.grey)
    pdf.drawString(40, 30, "Generated by Datadog AI-Powered Observability Reporting Module")
    pdf.save(); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{report_type}_datadog_report.pdf",
                     mimetype="application/pdf")


@app.route("/export/excel/<report_type>")
def export_excel(report_type):
    summary, reports = get_report_data(report_type)
    wb = Workbook(); ws = wb.active; ws.title = "Datadog Report"
    ws.merge_cells("A1:J1")
    ws["A1"] = "Datadog AI-Powered Observability Report"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="101935")
    ws["A1"].alignment = Alignment(horizontal="center")
    for row_i, (k, v) in enumerate([("Report Type", summary["title"]),
                                      ("Generated On", summary["generated_on"]),
                                      ("Average Uptime", f"{summary['avg_uptime']}%"),
                                      ("Total Alerts", summary["total_alerts"])], 3):
        ws.cell(row_i, 1, k); ws.cell(row_i, 2, v)

    hdrs = ["Date","Service Name","Status","Uptime %","Response (ms)",
            "Error Rate %","Alerts","CPU %","Memory %","AI Insight"]
    SR = 9
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(SR, ci, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3157D5")
        c.alignment = Alignment(horizontal="center")

    for ri, r in enumerate(reports, SR+1):
        for ci, v in enumerate([r["report_date"],r["service_name"],r["status"],
                                  r["uptime"],r["response_time"],r["error_rate"],
                                  r["alerts"],r["cpu_usage"],r["memory_usage"],
                                  r["ai_insight"]], 1):
            ws.cell(ri, ci, v).alignment = Alignment(vertical="center", wrap_text=True)

    for ci, w in enumerate([15,24,15,13,18,15,10,12,14,55], 1):
        ws.column_dimensions[chr(64+ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{report_type}_datadog_report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ══════════════════════════════════════════════════════
# INTEGRATION MODULE API
# ══════════════════════════════════════════════════════
@app.route("/api/metrics")
def live_metrics():
    cpu = psutil.cpu_percent(interval=0.1)
    mem = psutil.virtual_memory().percent
    return jsonify({
        "cpu":    cpu if cpu > 0 else random.randint(38, 46),
        "memory": mem,
        "status": "Connected"
    })


# ══════════════════════════════════════════════════════
if __name__ == "__main__":
    initialize_database()
    app.run(debug=True, port=5000)
